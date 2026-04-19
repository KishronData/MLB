"""
mlb_harvest.py
--------------
Pulls MLB game results and standings for the previous day and appends
them to Excel workbooks stored in OneDrive via the Microsoft Graph API.

Outputs two Excel files (configured in SETTINGS below):
  - GameResults.xlsx  : One row per team per completed game
  - Standings.xlsx    : Daily snapshot of division/league/overall rankings

Designed to run daily via GitHub Actions (see mlb_harvest.yml).

Dependencies:
    pip install MLB-StatsAPI openpyxl requests msal

Microsoft Graph / OneDrive setup:
    1. Register an app in Azure AD (Entra ID) with these API permissions:
         Files.ReadWrite (delegated or application)
    2. Create a client secret for the app registration
    3. Fill in CLIENT_ID, CLIENT_SECRET, TENANT_ID below (or use env vars)
"""

import os
import io
import logging
from datetime import date, timedelta

import statsapi          # MLB-StatsAPI wrapper: pip install MLB-StatsAPI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests
import msal              # Microsoft Authentication Library: pip install msal

# ---------------------------------------------------------------------------
# SETTINGS — Edit these each season or to point at different OneDrive folders
# ---------------------------------------------------------------------------

# The MLB season year to pull data for. Change this each new season.
SEASON_YEAR = 2026

# Only pull games of these types. "R" = Regular Season, "P" = Postseason.
# Excludes Spring Training ("S") and All-Star ("A").
VALID_GAME_TYPES = {"R", "P"}

# OneDrive folder path where Excel files will be stored.
# Use forward slashes. This folder must already exist in OneDrive.
# Example: "MLB Data/2025" or just "MLB Data"
ONEDRIVE_FOLDER = "Projects/MLB/Data"

# File names for each output workbook. Change to archive a prior season.
GAME_RESULTS_FILE = "GameResults_2026.xlsx"
STANDINGS_FILE    = "Standings_2026.xlsx"

# ---------------------------------------------------------------------------
# MICROSOFT GRAPH / AZURE AD CREDENTIALS
# Read from environment variables so secrets are never hardcoded in source.
# Set these as GitHub Actions secrets (or in your local .env for testing).
# ---------------------------------------------------------------------------

CLIENT_ID     = os.environ.get("AZURE_CLIENT_ID")      # Azure app registration client ID
CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET")  # Azure app registration client secret
TENANT_ID     = os.environ.get("AZURE_TENANT_ID")      # Azure AD tenant ID

# The OneDrive user whose drive will be written to.
# For a personal Microsoft account this is typically "me".
# For a work/school account you can use the UPN: "user@domain.com"
ONEDRIVE_USER = os.environ.get("ONEDRIVE_USER", "me")

# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# DATE HELPERS
# ---------------------------------------------------------------------------

def get_yesterday() -> date:
    """Return yesterday's date. This is the target date for each daily run."""
    return date.today() - timedelta(days=1)


def format_date_for_api(d: date) -> str:
    """Format a date as MM/DD/YYYY, which the MLB Stats API expects."""
    return d.strftime("%m/%d/%Y")


def format_date_iso(d: date) -> str:
    """Format a date as YYYY-MM-DD for storage in Excel."""
    return d.strftime("%Y-%m-%d")

# ---------------------------------------------------------------------------
# MLB DATA FETCHERS
# ---------------------------------------------------------------------------

def fetch_game_results(target_date: date) -> list[dict]:
    """
    Fetch completed game results for a given date.
    Calls the MLB Stats API directly via requests instead of the statsapi
    wrapper, which has inconsistent support for the hydrate parameter.
    """
    date_str = format_date_for_api(target_date)
    iso_str  = format_date_iso(target_date)

    log.info(f"Fetching game schedule for {date_str} ...")

    # Call the MLB Stats API directly — no wrapper quirks
    url = "https://statsapi.mlb.com/api/v1/schedule"
    params = {
        "sportId":   1,
        "date":      target_date.strftime("%Y-%m-%d"),  # API prefers YYYY-MM-DD
        "hydrate":   "linescore",
        "gameType":  "R,P",   # Regular Season and Postseason only
    }
    resp = requests.get(url, params=params, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    rows = []
    for game_date in data.get("dates", []):
        for game in game_date.get("games", []):

            # Skip games that haven't finished
            status = game.get("status", {}).get("detailedState", "")
            if status != "Final":
                log.debug(f"Skipping non-final game: {game.get('gamePk')} — {status}")
                continue

            # Skip if season year doesn't match
            if str(game.get("season", "")) != str(SEASON_YEAR):
                continue

            game_type   = game.get("gameType", "")
            game_number = game.get("gameNumber", 1)  # 1 or 2 for doubleheaders
            game_id     = game.get("gamePk", "")

            # Postseason series info (empty for regular season)
            series_desc  = game.get("seriesDescription", "")
            series_game  = game.get("seriesGameNumber", "")

            home         = game.get("teams", {}).get("home", {})
            away         = game.get("teams", {}).get("away", {})
            home_team    = home.get("team", {}).get("name", "")
            away_team    = away.get("team", {}).get("name", "")
            home_runs    = home.get("score", 0)
            away_runs    = away.get("score", 0)
            home_won     = home.get("isWinner", False)

            # --- Home team row ---
            rows.append({
                "GameDate":      iso_str,
                "Season":        SEASON_YEAR,
                "GameType":      game_type,
                "SeriesDesc":    series_desc,
                "SeriesGameNum": series_game,
                "GameNumber":    game_number,
                "GameID":        game_id,
                "Team":          home_team,
                "HomeAway":      "Home",
                "Opponent":      away_team,
                "RunsScored":    home_runs,
                "RunsAllowed":   away_runs,
                "Win":           1 if home_won else 0,
                "Loss":          0 if home_won else 1,
            })

            # --- Away team row ---
            rows.append({
                "GameDate":      iso_str,
                "Season":        SEASON_YEAR,
                "GameType":      game_type,
                "SeriesDesc":    series_desc,
                "SeriesGameNum": series_game,
                "GameNumber":    game_number,
                "GameID":        game_id,
                "Team":          away_team,
                "HomeAway":      "Away",
                "Opponent":      home_team,
                "RunsScored":    away_runs,
                "RunsAllowed":   home_runs,
                "Win":           0 if home_won else 1,
                "Loss":          1 if home_won else 0,
            })

    log.info(f"Found {len(rows) // 2} completed game(s) → {len(rows)} team-game rows.")
    return rows


def fetch_standings(target_date: date) -> list[dict]:
    """
    Fetch standings as of a given date for both the AL (103) and NL (104).

    Returns a list of row dicts, one per team, capturing their rank within
    their division, their league, and across all of MLB on that date.

    League IDs:
      103 = American League
      104 = National League
    """
    iso_str  = format_date_iso(target_date)
    date_str = format_date_for_api(target_date)

    log.info(f"Fetching standings for {date_str} ...")

    rows = []

    for league_id, league_name in [(103, "AL"), (104, "NL")]:
        # standings_data returns a list of division records
        divisions = statsapi.standings_data(
            leagueId=str(league_id),
            date=date_str,
            season=str(SEASON_YEAR),
        )

        # Build a flat list of all teams in this league with their division rank
        league_teams = []
        for div in divisions.values():
            div_name = div.get("div_name", "")
            for rank_in_div, team in enumerate(div.get("teams", []), start=1):
                league_teams.append({
                    "team_name":  team.get("name", ""),
                    "division":   div_name,
                    "league":     league_name,
                    "div_rank":   rank_in_div,
                    "wins":       team.get("w", 0),
                    "losses":     team.get("l", 0),
                    "pct":        team.get("pct", ".000"),
                    "gb":         team.get("gb", "-"),      # Games behind division leader
                    "league_gb":  team.get("league_gb", "-"),  # Games behind league leader
                })

        # Compute league rank by sorting on wins desc, losses asc within the league
        league_teams.sort(key=lambda t: (-int(t["wins"]), int(t["losses"])))
        for league_rank, team in enumerate(league_teams, start=1):
            team["league_rank"] = league_rank

        rows.extend(league_teams)

    # Compute overall MLB rank across all 30 teams
    rows.sort(key=lambda t: (-int(t["wins"]), int(t["losses"])))
    for mlb_rank, team in enumerate(rows, start=1):
        team["mlb_rank"] = mlb_rank

    # Attach the date to every row
    for row in rows:
        row["GameDate"] = iso_str
        row["Season"]   = SEASON_YEAR

    log.info(f"Fetched standings for {len(rows)} teams.")
    return rows

# ---------------------------------------------------------------------------
# EXCEL HELPERS
# ---------------------------------------------------------------------------

# Column definitions for GameResults sheet
GAME_RESULTS_COLUMNS = [
    "GameDate", "Season", "GameType", "SeriesDesc", "SeriesGameNum",
    "GameNumber", "GameID", "Team", "HomeAway", "Opponent",
    "RunsScored", "RunsAllowed", "Win", "Loss", "GameSeq",
]

# Column definitions for Standings sheet
STANDINGS_COLUMNS = [
    "GameDate", "Season", "Team", "Division", "League",
    "DivisionRank", "LeagueRank", "MLBRank",
    "Wins", "Losses", "WinPct", "GamesBack", "LeagueGamesBack",
]

# Map from row dict keys → Excel column headers (standings)
STANDINGS_KEY_MAP = {
    "GameDate":        "GameDate",
    "Season":          "Season",
    "team_name":       "Team",
    "division":        "Division",
    "league":          "League",
    "div_rank":        "DivisionRank",
    "league_rank":     "LeagueRank",
    "mlb_rank":        "MLBRank",
    "wins":            "Wins",
    "losses":          "Losses",
    "pct":             "WinPct",
    "gb":              "GamesBack",
    "league_gb":       "LeagueGamesBack",
}


def _make_header_style():
    """Return openpyxl style objects for the header row."""
    font   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    fill   = PatternFill("solid", start_color="17375E")  # Dark navy
    align  = Alignment(horizontal="center", vertical="center")
    return font, fill, align


def build_workbook_from_scratch(columns: list[str]) -> openpyxl.Workbook:
    """
    Create a brand-new workbook with a single formatted header row.
    Called only when the file doesn't yet exist in OneDrive.
    """
    wb     = openpyxl.Workbook()
    ws     = wb.active
    ws.title = "Data"

    font, fill, align = _make_header_style()
    for col_idx, col_name in enumerate(columns, start=1):
        cell            = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font       = font
        cell.fill       = fill
        cell.alignment  = align

    # Freeze the header row so it stays visible while scrolling
    ws.freeze_panes = "A2"

    return wb


def append_rows_to_workbook(wb: openpyxl.Workbook, data_rows: list[dict],
                            columns: list[str], key_map: dict | None = None) -> int:
    """
    Append data_rows to the 'Data' sheet of an existing workbook.
    Returns the number of rows appended.

    key_map: optional mapping from dict keys → column names, used when the
             dict keys differ from the column header names (standings data).
    """
    ws         = wb["Data"]
    next_row   = ws.max_row + 1
    row_font   = Font(name="Arial", size=10)

    for row_dict in data_rows:
        for col_idx, col_name in enumerate(columns, start=1):
            # Resolve the dict key: use key_map if provided, else col_name directly
            if key_map:
                dict_key = next(
                    (k for k, v in key_map.items() if v == col_name), col_name
                )
            else:
                dict_key = col_name

            value = row_dict.get(dict_key, "")
            cell  = ws.cell(row=next_row, column=col_idx, value=value)
            cell.font = row_font

        next_row += 1

    return len(data_rows)

# ---------------------------------------------------------------------------
# MICROSOFT GRAPH / ONEDRIVE INTEGRATION
# ---------------------------------------------------------------------------

def get_graph_token() -> str:
    """
    Obtain an OAuth2 access token for Microsoft Graph using the
    client credentials (app-only) flow. This doesn't require a user
    to be signed in, which is what we want for automated/scheduled runs.

    Required Azure AD app permissions:
      Files.ReadWrite.All (Application permission)
    """
    authority = f"https://login.microsoftonline.com/{TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        CLIENT_ID,
        authority=authority,
        client_credential=CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )
    if "access_token" not in result:
        raise RuntimeError(f"Failed to acquire Graph token: {result.get('error_description')}")
    return result["access_token"]


def get_onedrive_file(token: str, file_name: str) -> bytes | None:
    """
    Download a file from OneDrive. Returns raw bytes if found, None if not found.

    The file is looked up by path: ONEDRIVE_FOLDER / file_name
    """
    # Construct the Graph API path for the file
    # For personal accounts use /me/drive; for work accounts /users/{id}/drive
    drive_path = f"/me/drive" if ONEDRIVE_USER == "me" else f"/users/{ONEDRIVE_USER}/drive"
    url = (
        f"https://graph.microsoft.com/v1.0"
        f"{drive_path}/root:/{ONEDRIVE_FOLDER}/{file_name}:/content"
    )
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers)

    if resp.status_code == 404:
        log.info(f"'{file_name}' not found in OneDrive — will create new.")
        return None
    resp.raise_for_status()
    return resp.content


def put_onedrive_file(token: str, file_name: str, file_bytes: bytes) -> None:
    """
    Upload (create or overwrite) a file in OneDrive.
    Uses the simple upload endpoint (supports files up to 4 MB;
    for larger files the resumable upload API would be needed, but
    our Excel files will stay well under that threshold).
    """
    drive_path = f"/me/drive" if ONEDRIVE_USER == "me" else f"/users/{ONEDRIVE_USER}/drive"
    url = (
        f"https://graph.microsoft.com/v1.0"
        f"{drive_path}/root:/{ONEDRIVE_FOLDER}/{file_name}:/content"
    )
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    resp = requests.put(url, headers=headers, data=file_bytes)
    resp.raise_for_status()
    log.info(f"Uploaded '{file_name}' to OneDrive at '{ONEDRIVE_FOLDER}'.")

# ---------------------------------------------------------------------------
# MAIN ORCHESTRATION
# ---------------------------------------------------------------------------

def assign_game_sequence(new_rows: list[dict], existing_rows: list[dict]) -> list[dict]:
    """
    Assigns GameSeq to each new row — the cumulative number of games that
    team has played through that game, counting all prior history plus today.

    Works correctly for both daily runs and backfills because it reads the
    existing workbook rows first to get each team's current game count before
    incrementing for the new rows.

    Handles doubleheaders correctly because new_rows is already ordered by
    GameDate + GameNumber from fetch_game_results(), so game 1 of a
    doubleheader always gets a lower GameSeq than game 2.
    """
    from collections import defaultdict

    # Build a count of games already recorded per team from the existing workbook.
    # Each row in GameResults represents one team's appearance in one game,
    # so a simple row count per team gives us games played.
    game_counts = defaultdict(int)
    for row in existing_rows:
        game_counts[row["Team"]] += 1

    # Sort new rows by date then game number so doubleheader sequencing is correct
    new_rows_sorted = sorted(
        new_rows,
        key=lambda r: (r["GameDate"], r["GameNumber"])
    )

    # Assign the next sequence number for each team
    for row in new_rows_sorted:
        team = row["Team"]
        game_counts[team] += 1
        row["GameSeq"] = game_counts[team]

    return new_rows_sorted


def read_rows_from_workbook(wb: openpyxl.Workbook, columns: list[str]) -> list[dict]:
    """
    Reads all data rows from the 'Data' sheet of an existing workbook and
    returns them as a list of dicts keyed by column name.

    Row 1 is always the header row and is skipped.
    Used by process_file to read existing GameResults rows before appending,
    so assign_game_sequence can correctly compute GameSeq.
    """
    ws   = wb["Data"]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip completely empty rows (can appear at end of file)
        if not any(row):
            continue
        row_dict = {
            col: row[idx]
            for idx, col in enumerate(columns)
            if idx < len(row)
        }
        rows.append(row_dict)

    return rows


def process_file(token: str, file_name: str, new_rows: list[dict],
                 columns: list[str], key_map: dict | None = None) -> None:
    """
    Full round-trip for one Excel file:
      1. Try to download existing file from OneDrive
      2. If not found, create a new workbook with headers
      3. For GameResults only: read existing rows back out so assign_game_sequence
         can compute the correct GameSeq values before appending
      4. Append today's new rows
      5. Upload the updated workbook back to OneDrive

    The key_map parameter is only used for Standings, where the dict keys
    returned by fetch_standings() differ from the Excel column header names.
    GameResults dict keys match column names directly so key_map is None.
    """
    if not new_rows:
        log.warning(f"No new rows to append to '{file_name}'. Skipping upload.")
        return

    # Step 1: Download existing workbook (or start fresh)
    existing_bytes = get_onedrive_file(token, file_name)
    if existing_bytes:
        wb = openpyxl.load_workbook(io.BytesIO(existing_bytes))
        log.info(f"Loaded existing '{file_name}' from OneDrive.")
    else:
        wb = build_workbook_from_scratch(columns)
        log.info(f"Created new workbook for '{file_name}'.")

    # Step 2: For GameResults, read existing rows back out of the workbook
    # so assign_game_sequence knows each team's current game count.
    # Standings doesn't need this so we skip it when key_map is present.
    if key_map is None:
        existing_rows = read_rows_from_workbook(wb, columns)
        log.info(f"Read {len(existing_rows)} existing rows from '{file_name}'.")
        new_rows = assign_game_sequence(new_rows, existing_rows)
        log.info(f"GameSeq assigned to {len(new_rows)} new rows.")

    # Step 3: Append new rows
    n = append_rows_to_workbook(wb, new_rows, columns, key_map)
    log.info(f"Appended {n} rows to '{file_name}'.")

    # Step 4: Serialize workbook to bytes and upload
    buf = io.BytesIO()
    wb.save(buf)
    put_onedrive_file(token, file_name, buf.getvalue())


# MAIN that gets game data from yesterday
def main():
    target_date = get_yesterday()
    log.info(f"=== MLB Harvest starting for {target_date} ===")

    # Fetch MLB data
    game_rows     = fetch_game_results(target_date)
    standing_rows = fetch_standings(target_date)

    # Authenticate with Microsoft Graph
    log.info("Authenticating with Microsoft Graph ...")
    token = get_graph_token()

    # Process and upload GameResults
    process_file(
        token       = token,
        file_name   = GAME_RESULTS_FILE,
        new_rows    = game_rows,
        columns     = GAME_RESULTS_COLUMNS,
        key_map     = None,   # dict keys already match column names
    )

    # Process and upload Standings
    process_file(
        token       = token,
        file_name   = STANDINGS_FILE,
        new_rows    = standing_rows,
        columns     = STANDINGS_COLUMNS,
        key_map     = STANDINGS_KEY_MAP,
    )

    log.info("=== MLB Harvest complete ===")


# MAIN that backfills data as needed
#from datetime import date, timedelta

#def main():
#    start = date(2026, 4, 18)   # Opening Day 2025
#    end   = date(2026, 4, 18)   # End of regular season (adjust as needed)
#    d     = start
#    token = get_graph_token()
#    while d <= end:
#        game_rows     = fetch_game_results(d)
#        standing_rows = fetch_standings(d)
#        process_file(token, GAME_RESULTS_FILE, game_rows,     GAME_RESULTS_COLUMNS)
#        process_file(token, STANDINGS_FILE,    standing_rows, STANDINGS_COLUMNS, STANDINGS_KEY_MAP)
#        d += timedelta(days=1)
  

if __name__ == "__main__":
    main()
